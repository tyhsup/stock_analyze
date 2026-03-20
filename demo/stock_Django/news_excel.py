"""
news_excel.py — Per-stock, per-year Excel news storage manager.

Folder structure:
    demo/newsapp/news_data/
        AAPL.xlsx          (sheets: "2024", "2025", ...)
        2330.xlsx          (sheets: "2024", "2025", ...)
        _general.xlsx      (general/unbounded news)

Schema per sheet (columns):
    標題 | 日期 | 內容 | 連結 | 正負分析 | 來源
"""

import os
import logging
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import threading

logger = logging.getLogger(__name__)

# Absolute path to news data folder - FIXED as per user request
NEWS_DATA_DIR = Path(r'E:\Infinity\mydjango\demo\newsapp\news_data')

SCHEMA_COLUMNS = ['標題', '日期', '內容', '連結', '正負分析', '來源']

# Threading lock for file operations
_excel_lock = threading.Lock()


class NewsExcelManager:
    """Manages per-stock, per-year news stored in Excel files."""

    def __init__(self, data_dir: Path = None):
        self.data_dir = data_dir or NEWS_DATA_DIR
        self.data_dir.mkdir(parents=True, exist_ok=True)

    def get_file_path(self, ticker: str) -> Path:
        """Return the Excel file path for a given ticker."""
        # Sanitize ticker for filename (e.g., 2330.TW → 2330_TW.xlsx)
        safe_name = ticker.upper().replace('.', '_')
        return self.data_dir / f"{safe_name}.xlsx"

    def _get_year_from_row(self, row: dict) -> str:
        """Extract year string from a news row dict."""
        date_val = row.get('日期', '')
        try:
            if isinstance(date_val, datetime):
                return str(date_val.year)
            dt = pd.to_datetime(str(date_val), errors='coerce')
            if pd.notna(dt):
                return str(dt.year)
        except Exception:
            pass
        return str(datetime.now().year)

    def read_news(self, ticker: str, query: str = None, years: list = None, limit: int = 1000) -> list:
        """
        Read news for a ticker, optionally filtered by keyword and year.
        Searches across ALL sheets (years) unless years is specified.
        Older data files (newsapp/static/news_data.xlsx) are NOT affected.
        """
        file_path = self.get_file_path(ticker)

        # Also check legacy general file for backwards compatibility
        legacy_path = Path(__file__).resolve().parent.parent / 'newsapp' / 'static' / 'news_data.xlsx'

        frames = []

        # Read from per-stock file
        if file_path.exists():
            try:
                xl = pd.ExcelFile(file_path)
                sheets_to_read = years if years else xl.sheet_names
                for sheet in sheets_to_read:
                    if sheet in xl.sheet_names:
                        df = xl.parse(sheet)
                        if not df.empty:
                            frames.append(df)
            except Exception as e:
                logger.warning(f"Error reading {file_path}: {e}")

    def _normalize_df_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize DataFrame columns to SCHEMA_COLUMNS.
        Handles cases where headers might be mangled or missing '內容'.
        """
        if df.empty:
            return pd.DataFrame(columns=SCHEMA_COLUMNS)
            
        # Try to detect if this is a 5-column (old) or 6-column (new) file
        cols = list(df.columns)
        if len(cols) == 5:
            # Old schema: 標題, 日期, 連結, 正負分析, 來源
            df.columns = ['標題', '日期', '連結', '正負分析', '來源']
            df['內容'] = ''
        elif len(cols) == 6:
            # New schema: 標題, 日期, 內容, 連結, 正負分析, 來源
            df.columns = SCHEMA_COLUMNS
        else:
            # Unknown schema, try to map available ones, fill missing
            for col in SCHEMA_COLUMNS:
                if col not in df.columns:
                    df[col] = ''
        
        return df[SCHEMA_COLUMNS]

    def read_news(self, ticker: str, query: str = None, limit: int = 100) -> list:
        """Read news for a ticker and filter by query."""
        file_path = self.get_file_path(ticker)
        if not file_path.exists():
            return []

        all_frames = []
        try:
            xl = pd.ExcelFile(file_path)
            for sheet in xl.sheet_names:
                df = xl.parse(sheet)
                if not df.empty:
                    all_frames.append(self._normalize_df_columns(df))
        except Exception as e:
            logger.error(f"Error reading {file_path}: {e}")
            return []

        if not all_frames:
            return []

        combined = pd.concat(all_frames, ignore_index=True)

        # Filter by keyword (in title or content)
        if query:
            mask = (
                combined['標題'].astype(str).str.contains(query, case=False, na=False) |
                combined['內容'].astype(str).str.contains(query, case=False, na=False)
            )
            combined = combined[mask]

        # Sort by date descending
        combined['日期'] = pd.to_datetime(combined['日期'], errors='coerce')
        combined = combined.sort_values('日期', ascending=False)
        combined['日期'] = combined['日期'].dt.strftime('%Y-%m-%d').fillna('')

        # Fill NaN with empty string to avoid JSON errors in frontend
        return combined[SCHEMA_COLUMNS].head(limit).fillna('').to_dict(orient='records')

    def read_news_general(self, query: str = None, limit: int = 1000) -> list:
        """Search across all tickers' Excel files."""
        all_frames = []
        for xlsx_file in self.data_dir.glob('*.xlsx'):
            try:
                xl = pd.ExcelFile(xlsx_file)
                for sheet in xl.sheet_names:
                    df = xl.parse(sheet)
                    if not df.empty:
                        df = self._normalize_df_columns(df)
                        if '來源' not in df.columns or not df['來源'].any():
                            df['來源'] = xlsx_file.stem
                        all_frames.append(df)
            except Exception:
                pass

        if not all_frames:
            return []

        combined = pd.concat(all_frames, ignore_index=True)

        if query:
            mask = (
                combined['標題'].astype(str).str.contains(query, case=False, na=False) |
                combined['內容'].astype(str).str.contains(query, case=False, na=False)
            )
            combined = combined[mask]

        combined['日期'] = pd.to_datetime(combined['日期'], errors='coerce')
        combined = combined.sort_values('日期', ascending=False)
        combined['日期'] = combined['日期'].dt.strftime('%Y-%m-%d').fillna('')

        return combined[SCHEMA_COLUMNS].head(limit).fillna('').to_dict(orient='records')

    def write_news(self, ticker: str, news_list: list) -> int:
        """
        Write news to the ticker's Excel file, separated by year in different sheets.
        Deduplicates by (標題 + 日期) combination.
        Returns the number of new rows added.
        """
        if not news_list:
            return 0

        file_path = self.get_file_path(ticker)

        with _excel_lock:
            # Group incoming news by year
            year_groups: dict[str, list] = {}
            for item in news_list:
                year = self._get_year_from_row(item)
                year_groups.setdefault(year, []).append(item)

            total_added = 0

            # Load or create workbook
            if file_path.exists():
                wb = load_workbook(file_path)
            else:
                wb = Workbook()
                # Remove default empty sheet
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']

            for year, rows in year_groups.items():
                new_df = pd.DataFrame(rows)
                for col in SCHEMA_COLUMNS:
                    if col not in new_df.columns:
                        new_df[col] = ''
                new_df = new_df[SCHEMA_COLUMNS]
                new_df['日期'] = pd.to_datetime(new_df['日期'], errors='coerce')
                new_df['日期'] = new_df['日期'].dt.strftime('%Y-%m-%d')

                if year in wb.sheetnames:
                    # Read existing, deduplicate
                    ws = wb[year]
                    existing_data = []
                    headers = None
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i == 0:
                            headers = list(row)
                        else:
                            if headers:
                                existing_data.append(dict(zip(headers, row)))
                    existing_df = self._normalize_df_columns(pd.DataFrame(existing_data)) if existing_data else pd.DataFrame(columns=SCHEMA_COLUMNS)
                else:
                    existing_df = pd.DataFrame(columns=SCHEMA_COLUMNS)
                    wb.create_sheet(title=year)

                # Dedup key = (標題, 日期) or (連結)
                if not existing_df.empty:
                    # Multi-key deduplication
                    existing_keys = set(zip(
                        existing_df['標題'].astype(str),
                        existing_df['日期'].astype(str)
                    ))
                    existing_links = set(existing_df['連結'].astype(str).tolist())
                    
                    new_rows = new_df[~new_df.apply(
                        lambda r: (str(r['標題']), str(r['日期'])) in existing_keys or str(r['連結']) in existing_links, axis=1
                    )]
                else:
                    new_rows = new_df

                if new_rows.empty:
                    continue

                combined = pd.concat([existing_df, new_rows], ignore_index=True)
                combined['日期'] = pd.to_datetime(combined['日期'], errors='coerce')
                combined = combined.sort_values('日期', ascending=False)
                combined['日期'] = combined['日期'].dt.strftime('%Y-%m-%d').fillna('')

                # Rewrite sheet
                ws = wb[year]
                ws.delete_rows(1, ws.max_row)
                for r in dataframe_to_rows(combined[SCHEMA_COLUMNS], index=False, header=True):
                    ws.append(r)

                total_added += len(new_rows)

            wb.save(file_path)
            logger.info(f"[NewsExcel] {ticker}: written {total_added} new rows to {file_path}")
            return total_added

    def get_latest_date(self, ticker: str) -> str | None:
        """Return the latest news date string for a ticker, or None if no data."""
        file_path = self.get_file_path(ticker)
        if not file_path.exists():
            return None
        try:
            xl = pd.ExcelFile(file_path)
            dates = []
            for sheet in xl.sheet_names:
                df = xl.parse(sheet, usecols=['日期'] if '日期' in xl.parse(sheet, nrows=1).columns else None)
                if '日期' in df.columns:
                    d = pd.to_datetime(df['日期'], errors='coerce').dropna()
                    if not d.empty:
                        dates.append(d.max())
            if dates:
                return str(max(dates).strftime('%Y-%m-%d'))
        except Exception as e:
            logger.warning(f"Cannot get latest_date for {ticker}: {e}")
        return None
