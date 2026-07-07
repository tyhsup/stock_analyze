import re
import openpyxl
from openpyxl.utils import get_column_letter

class ExcelModelAuditor:
    @staticmethod
    def audit_workbook(file_stream) -> dict:
        """
        對上傳的 Excel 檔案進行公式與數據稽核，回傳詳細的問題報告。
        Traditional Chinese: 財務模型稽核引擎，分析公式、偵測硬編碼、核對勾稽關係與孤立單元格。
        """
        try:
            # 1. 載入含有公式的 workbook (data_only=False)
            wb_formula = openpyxl.load_workbook(file_stream, data_only=False)
            
            # 2. 重新定位流，載入僅有計算值的 workbook (data_only=True)
            file_stream.seek(0)
            wb_value = openpyxl.load_workbook(file_stream, data_only=True)
            
            issues = []
            sheets_audited = []
            
            # 分析工作表中所有公式對單元格的參照，用以尋找 Orphan Cells (孤立參數單元格)
            all_formula_refs = set()
            cell_refs_pattern = re.compile(r'\b([A-Z]+[0-9]+)\b')
            
            # 蒐集所有公式所引用的單元格
            for sheet_name in wb_formula.sheetnames:
                ws_f = wb_formula[sheet_name]
                for r in range(1, ws_f.max_row + 1):
                    for c in range(1, ws_f.max_column + 1):
                        cell_val = ws_f.cell(row=r, column=c).value
                        if isinstance(cell_val, str) and cell_val.startswith('='):
                            # 提取所有單元格引用 (如 A1, B2)
                            refs = cell_refs_pattern.findall(cell_val.upper())
                            for ref in refs:
                                all_formula_refs.add((sheet_name, ref))
            
            # 遍歷各個工作表進行細化稽核
            for sheet_name in wb_formula.sheetnames:
                ws_f = wb_formula[sheet_name]
                ws_v = wb_value[sheet_name]
                sheets_audited.append(sheet_name)
                
                # 財務指標單元格追蹤 (用於勾稽校驗)
                ev_val, net_debt_val, equity_val = None, None, None
                ev_cell, net_debt_cell, equity_cell = None, None, None
                
                for r in range(1, ws_f.max_row + 1):
                    for c in range(1, ws_f.max_column + 1):
                        cell_coord = f"{get_column_letter(c)}{r}"
                        cell_f = ws_f.cell(row=r, column=c)
                        cell_v = ws_v.cell(row=r, column=c)
                        
                        val_f = cell_f.value
                        val_v = cell_v.value
                        
                        # 1. 斷鏈與公式錯誤偵測
                        err_keywords = ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A', '#NUM!']
                        if val_f and isinstance(val_f, str) and any(kw in val_f.upper() for kw in err_keywords):
                            issues.append({
                                "sheet": sheet_name,
                                "cell": cell_coord,
                                "type": "broken_formula",
                                "severity": "high",
                                "message": f"公式包含損壞的參照或語法錯誤: '{val_f}'",
                                "suggestion": "檢查公式所引用的單元格是否已被刪除或重新命名。"
                            })
                            continue
                            
                        if val_v and isinstance(val_v, str) and any(kw in val_v.upper() for kw in err_keywords):
                            issues.append({
                                "sheet": sheet_name,
                                "cell": cell_coord,
                                "type": "calculated_error",
                                "severity": "high",
                                "message": f"單元格計算出錯誤值: '{val_v}' (公式: '{val_f}')",
                                "suggestion": "檢查除數是否為零，或參照的單元格內容類型是否不一致。"
                            })
                            continue
                        
                        # 2. 硬編碼數值偵測 (Hardcoded Value in Projection Cells)
                        if r > 2 and c > 1 and val_v is not None and not isinstance(val_f, str) and isinstance(val_v, (int, float)):
                            row_header = str(ws_f.cell(row=r, column=1).value or '')
                            col_header = str(ws_f.cell(row=2, column=c).value or '')
                            keyword_match = any(kw in row_header.lower() or kw in col_header.lower() for kw in ['revenue', 'fcf', 'ebit', 'tax', 'discount', 'implied', 'fair value', '營收', '自由現金', '折現', '估值'])
                            
                            if keyword_match:
                                issues.append({
                                    "sheet": sheet_name,
                                    "cell": cell_coord,
                                    "type": "hardcoded_value",
                                    "severity": "medium",
                                    "message": f"預測或估值區域檢測到硬編碼常數: {val_v} (標題: '{row_header}')",
                                    "suggestion": "建議將此硬編碼數值改用動態公式鏈結，以提高財務模型的靈活性。"
                                })
                        
                        # 3. 收集特定財務指標以進行勾稽關係分析
                        label = str(ws_f.cell(row=r, column=1).value or '')
                        if 'enterprise value' in label.lower() or '企業價值' in label:
                            ev_val = val_v
                            ev_cell = cell_coord
                        elif 'net debt' in label.lower() or '淨負債' in label:
                            net_debt_val = val_v
                            net_debt_cell = cell_coord
                        elif 'equity value' in label.lower() or '股權價值' in label:
                            equity_val = val_v
                            equity_cell = cell_coord
                            
                        # 4. 檢測未參照引用的核心參數單元格 (Orphan Parameter Cells)
                        if 'wacc' in label.lower() or '折現率' in label or 'growth rate' in label.lower() or '成長率' in label:
                            if (sheet_name, cell_coord) not in all_formula_refs and isinstance(val_v, (int, float)):
                                issues.append({
                                    "sheet": sheet_name,
                                    "cell": cell_coord,
                                    "type": "orphan_parameter",
                                    "severity": "low",
                                    "message": f"核心參數單元格 '{label}' (值: {val_v}) 未被任何公式引用",
                                    "suggestion": "請確認該參數是否有正確關聯至 DCF 或是估值公式中，否則修改此參數將無法動態重算。"
                                })
                
                # 5. 勾稽關係一致性校驗 (EV - Net Debt == Equity Value)
                if ev_val is not None and net_debt_val is not None and equity_val is not None:
                    try:
                        expected_equity = float(ev_val) - float(net_debt_val)
                        actual_equity = float(equity_val)
                        if abs(expected_equity - actual_equity) > 0.05: # 允許 0.05 誤差範圍
                            issues.append({
                                "sheet": sheet_name,
                                "cell": equity_cell,
                                "type": "mismatch_error",
                                "severity": "high",
                                "message": f"勾稽一致性失敗: 股權價值 ({actual_equity}) 不等於 企業價值 ({ev_val}) - 淨負債 ({net_debt_val})",
                                "suggestion": f"請檢查 {equity_cell}、{ev_cell} 及 {net_debt_cell} 單元格的公式勾稽鏈結關係是否正確。"
                            })
                    except Exception:
                        pass
                        
            return {
                "status": "success",
                "sheets_audited": sheets_audited,
                "total_issues": len(issues),
                "issues": issues
            }
            
        except Exception as e:
            return {
                "status": "error",
                "message": f"Excel 模型稽核解析異常: {str(e)}"
            }
