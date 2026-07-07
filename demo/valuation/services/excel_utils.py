import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def generate_valuation_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Model"
    
    # 啟用格線
    ws.views.sheetView[0].showGridLines = True
    
    # 樣式定義
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    
    fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_accent = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    border_double_bottom = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # 1. 標題
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Financial Valuation Model: {data.get('symbol', 'Target')}"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # 2. 假設參數區 (B 欄為輸入區)
    ws["A3"] = "Key Assumptions & Parameters"
    ws["A3"].font = font_section
    
    assumptions_labels = [
        ("WACC (折現率)", data.get("assumptions", {}).get("wacc", 0.08), "0.0%"),
        ("Perpetuity Growth Rate (永續成長率 g)", data.get("assumptions", {}).get("exit_growth_rate", 0.02), "0.0%"),
        ("Tax Rate (所得稅率)", data.get("assumptions", {}).get("tax_rate", 0.20), "0.0%"),
        ("Current Share Price (目前股價)", data.get("current_price", 0.0), "$#,##0.00"),
        ("Shares Outstanding (M) (發行股數)", data.get("dcf", {}).get("shares_outstanding", 1.0) / 1000000.0, "#,##0.00"),
        ("Net Debt (M) (淨負債)", data.get("dcf", {}).get("net_debt", 0.0) / 1000000.0, "$#,##0.00"),
        ("DCF Method Weight", data.get("dcf_weight_pct", 50) / 100.0, "0%"),
        ("Market Method Weight", data.get("market_weight_pct", 50) / 100.0, "0%"),
        ("Sentiment Premium (情緒溢價)", 1.0, "0.0%"),
    ]
    
    for idx, (label, val, fmt) in enumerate(assumptions_labels):
        row = 4 + idx
        ws.cell(row=row, column=1, value=label).font = font_regular
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.font = font_bold
        cell_val.number_format = fmt
        cell_val.alignment = Alignment(horizontal="right")
        cell_val.fill = fill_light
        ws.cell(row=row, column=1).border = border_thin
        cell_val.border = border_thin
        
    # 3. 預測投影區 (5年預測)
    ws["A15"] = "5-Year Financial Projections (in Millions)"
    ws["A15"].font = font_section
    
    headers = ["Metric (M)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for col_idx, h in enumerate(headers):
        cell = ws.cell(row=16, column=col_idx + 1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        
    # 填入營收與自由現金流的基數
    dcf_data = data.get("dcf", {})
    proj_fcf = dcf_data.get("projected_fcf", {})
    revenues = proj_fcf.get("revenues", [])
    fcfs = proj_fcf.get("fcfs", [])
    
    # 營收列 (第 17 列)
    ws.cell(row=17, column=1, value="Revenue (營業收入)").font = font_regular
    # 自由現金流列 (第 18 列)
    ws.cell(row=18, column=1, value="Free Cash Flow (自由現金流 FCF)").font = font_regular
    # 折現因子列 (第 19 列) - 使用公式：=1 / ((1 + $B$4)^t) (B4 是 WACC)
    ws.cell(row=19, column=1, value="Discount Factor (折現因子)").font = font_regular
    # 折現自由現金流列 (第 20 列) - 使用公式：=FCF * 折現因子
    ws.cell(row=20, column=1, value="Discounted FCF (折現現金流)").font = font_regular
    
    for i in range(5):
        c = 2 + i
        rev_val = revenues[i] if i < len(revenues) else 0.0
        fcf_val = fcfs[i] if i < len(fcfs) else 0.0
        
        ws.cell(row=17, column=c, value=rev_val).number_format = "$#,##0.00"
        ws.cell(row=18, column=c, value=fcf_val).number_format = "$#,##0.00"
        
        # 折現因子公式：=1 / ((1 + $B$4)^{i+1}) (B4 是 WACC)
        ws.cell(row=19, column=c, value=f"=1 / ((1 + $B$4)^{i+1})").number_format = "0.0000"
        
        # 折現 FCF 公式：=C18 * C19
        col_letter = get_column_letter(c)
        ws.cell(row=20, column=c, value=f"={col_letter}18 * {col_letter}19").number_format = "$#,##0.00"
        
        for r in range(17, 21):
            ws.cell(row=r, column=c).font = font_regular if r != 20 else font_bold
            ws.cell(row=r, column=c).border = border_thin
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
            
    # 4. DCF 估值計算結果區 (使用動態公式)
    ws["A22"] = "DCF Valuation Summary"
    ws["A22"].font = font_section
    
    # 終值 (Terminal Value) 公式：=F18 * (1 + B5) / (B4 - B5) (F18 是第 5 年的 FCF，B5 是永續成長率，B4 是 WACC)
    # 注意在 Excel 表中，WACC 是在 B4，永續成長率是在 B5，所得稅率是在 B6，目前股價在 B7，發行股數在 B8，淨負債在 B9，DCF權重在 B10，乘數權重在 B11，情緒溢價在 B12。
    ws.cell(row=23, column=1, value="Terminal Value (終值)").font = font_regular
    ws.cell(row=23, column=2, value="=F18 * (1 + B5) / (B4 - B5)").number_format = "$#,##0.00"
    
    # 終值折現值 (PV of TV) 公式：=B23 * F19 (F19 是第 5 年的折現因子)
    ws.cell(row=24, column=1, value="PV of Terminal Value (終值折現值)").font = font_regular
    ws.cell(row=24, column=2, value="=B23 * F19").number_format = "$#,##0.00"
    
    # 累計預測 FCF 折現值 (PV of FCFs) 公式：=SUM(B20:F20)
    ws.cell(row=25, column=1, value="PV of Projected FCFs (預測期現金流折現值)").font = font_regular
    ws.cell(row=25, column=2, value="=SUM(B20:F20)").number_format = "$#,##0.00"
    
    # 企業價值 (Enterprise Value) 公式：=B24 + B25
    ws.cell(row=26, column=1, value="Enterprise Value (企業價值 EV)").font = font_regular
    ws.cell(row=26, column=2, value="=B24 + B25").number_format = "$#,##0.00"
    
    # 淨負債 (引自 B9)
    ws.cell(row=27, column=1, value="Less: Net Debt (減: 淨負債)").font = font_regular
    ws.cell(row=27, column=2, value="=B9").number_format = "$#,##0.00"
    
    # 股權價值 (Equity Value) 公式：=B26 - B27
    ws.cell(row=28, column=1, value="Equity Value (股權價值)").font = font_bold
    ws.cell(row=28, column=2, value="=B26 - B27").number_format = "$#,##0.00"
    ws.cell(row=28, column=2).fill = fill_accent
    
    # 發行股數 (引自 B8)
    ws.cell(row=29, column=1, value="Shares Outstanding (發行股數)").font = font_regular
    ws.cell(row=29, column=2, value="=B8").number_format = "#,##0.00"
    
    # DCF 每股合理價值 (DCF Implied Price) 公式：=B28 / B29
    ws.cell(row=30, column=1, value="DCF Implied Share Price (DCF 每股估計價值)").font = font_bold
    ws.cell(row=30, column=2, value="=B28 / B29").number_format = "$#,##0.00"
    ws.cell(row=30, column=2).fill = fill_accent
    
    # 5. 乘數估值法 (Market Approach)
    ws["D22"] = "Relative Valuation Summary"
    ws["D22"].font = font_section
    
    market_data = data.get("market_approach", {})
    ws.cell(row=23, column=4, value="PE Implied Price (PE 乘數估計值)").font = font_regular
    ws.cell(row=23, column=5, value=market_data.get("pe_price", 0.0)).number_format = "$#,##0.00"
    
    ws.cell(row=24, column=4, value="EV/EBITDA Implied Price (EV/EBITDA 估計值)").font = font_regular
    ws.cell(row=24, column=5, value=market_data.get("ev_ebitda_price", 0.0)).number_format = "$#,##0.00"
    
    # 乘數平均估計值 公式：=AVERAGE(E23:E24)
    ws.cell(row=25, column=4, value="Market Implied Share Price (乘數平均價值)").font = font_bold
    ws.cell(row=25, column=5, value="=AVERAGE(E23:E24)").number_format = "$#,##0.00"
    ws.cell(row=25, column=5).fill = fill_accent
    
    # 6. 最終綜合權重合理價值 (Blended Fair Value)
    ws["A32"] = "Final Blended Fair Value Calculation"
    ws["A32"].font = font_section
    
    # 綜合價值公式：=((B30 * B10) + (E25 * B11)) * B12 (B30 是 DCF Implied Price，B10 是 DCF Weight，E25 是 Market Implied Price，B11 是 Market Weight，B12 是 Sentiment Premium)
    ws.cell(row=33, column=1, value="Blended Intrinsic Value (綜合合理價值)").font = font_bold
    ws.cell(row=33, column=2, value="=((B30 * B10) + (E25 * B11)) * B12").number_format = "$#,##0.00"
    ws.cell(row=33, column=2).fill = fill_accent
    ws.cell(row=33, column=2).border = border_double_bottom
    
    # 目前股價 (引自 B7)
    ws.cell(row=34, column=1, value="Current Share Price (目前市場價格)").font = font_regular
    ws.cell(row=34, column=2, value="=B7").number_format = "$#,##0.00"
    
    # 潛在漲跌幅 公式：=(B33 / B34) - 1
    ws.cell(row=35, column=1, value="Upside / Downside (潛在漲跌幅)").font = font_bold
    ws.cell(row=35, column=2, value="=(B33 / B34) - 1").number_format = "0.0%"
    ws.cell(row=35, column=2).border = border_double_bottom
    
    # 美化所有結果區儲存格
    for r in range(23, 31):
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=2).border = border_thin
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
        
        ws.cell(row=r, column=4).border = border_thin if r <= 25 else None
        ws.cell(row=r, column=5).border = border_thin if r <= 25 else None
        if r <= 25:
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="right")
            
    for r in range(33, 36):
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=2).border = border_thin if r == 34 else border_double_bottom
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
        
    # 自動調整欄寬
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # 特別拉寬第一欄，因為中文字較長
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['D'].width = 35
    
    # 將 workbook 寫入記憶體流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
